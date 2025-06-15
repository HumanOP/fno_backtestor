import os
import re
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from itertools import product, repeat
import itertools
import datetime
from openpyxl.drawing.image import Image
import json
from functools import partial, lru_cache
from backtesting_opt1 import Backtest
from typing import Union, Tuple, Callable
from tqdm import tqdm as _tqdm

class HyperParameterOptimizer:
    def __init__(self, db_path: str, strategy, *, cash: float = 100000, commission_per_contract: float = 0.65, 
                 option_multiplier: int = 75, legs: dict):
        self._backtest_factory = partial(
            Backtest, db_path=db_path, strategy=strategy, cash=cash,
            commission_per_contract=commission_per_contract, option_multiplier=option_multiplier
        )
        self.legs = legs

    def optimize(self, hyperparameter_grid: dict, maximize: str = 'Sharpe Ratio', 
                 method: str = 'grid', start_date: str = None, end_date: str = None,
                 max_tries: int = None, constraint: callable = None, random_state: int = None,
                 return_heatmap: bool = False, return_optimization: bool = False):
        """
        Optimize hyperparameters using grid search or SAMBO (Bayesian optimization).
        
        Parameters:
        - hyperparameter_grid: dict, parameter combinations to test
        - maximize: str, metric to maximize (default: 'Sharpe Ratio')
        - method: str, optimization method ('grid' or 'sambo')
        - max_tries: int, maximum parameter combinations to test
        - constraint: callable, function to filter parameter combinations
        - start_date: str, start date for backtest window (YYYY-MM-DD format)
        - end_date: str, end date for backtest window (YYYY-MM-DD format)
        - random_state: int, random seed for reproducible results (SAMBO only)
        - return_heatmap: bool, whether to return optimization heatmap (SAMBO only)
        - return_optimization: bool, whether to return optimization object (SAMBO only)
        
        Returns:
        - best_params: dict, best parameter combination
        - best_score: float, best metric value
        - results_df: pd.DataFrame, all results (grid search only)
        """
        if method not in ['grid', 'sambo']:
            raise ValueError(f"Method should be 'grid' or 'sambo', not {method!r}")
            
        # Log the optimization window if specified
        if start_date or end_date:
            print(f"Optimizing on date window: {start_date} to {end_date}")
        else:
            print("Optimizing on full dataset")

        if method == 'grid':
            return self._optimize_grid(hyperparameter_grid, maximize, start_date, end_date, max_tries, constraint)
        elif method == 'sambo':
            return self._optimize_sambo(hyperparameter_grid, maximize, start_date, end_date, max_tries, constraint, 
                                      random_state, return_heatmap, return_optimization)

    def _optimize_grid(self, hyperparameter_grid: dict, maximize: str, start_date: str, end_date: str, 
                      max_tries: int, constraint: callable):
        """Original grid search implementation"""
        param_keys = list(hyperparameter_grid.keys())
        param_values = [hyperparameter_grid[key] for key in param_keys]
        param_combinations = [dict(zip(param_keys, combo)) for combo in product(*param_values)]

        if constraint:
            param_combinations = [params for params in param_combinations if constraint(params)]
        if max_tries and max_tries < len(param_combinations):
            param_combinations = param_combinations[:max_tries]
        if not param_combinations:
            raise ValueError("No admissible parameter combinations to test")
            
        print(f"Total parameter combinations to test: {len(param_combinations)}")
        results = []
        best_sharpe = -np.inf
        best_params = None

        for idx, flat_params in enumerate(param_combinations, 1):
            print(f"Testing parameters {idx}/{len(param_combinations)}: {flat_params}")

            # Construct the full parameter dictionary
            params = {
                "iv_slope_thresholds": {
                    "upper_gamma": flat_params["upper_gamma"],
                    "upper_buffer": flat_params["upper_buffer"],
                    "lower_buffer": flat_params["lower_buffer"],
                    "lower_gamma": flat_params["lower_gamma"]
                },
                "portfolio_sl": flat_params.get("portfolio_sl", 0.01),  # Default if not optimized
                "portfolio_tp": flat_params.get("portfolio_tp", 0.03),  # Default if not optimized
                "legs": self.legs  # Pass legs as a parameter
            }

            try:
                backtest = self._backtest_factory()
                
                # Use run_window if date range is specified, otherwise use regular run
                if start_date is not None or end_date is not None:
                    result = backtest.run_window(start_date=start_date, end_date=end_date, **params)
                else:
                    result = backtest.run(**params)
                    
                print(f"Results: {result}")

                if maximize not in result or pd.isna(result[maximize]):
                    print(f"Invalid result for parameters {flat_params}: missing or NaN {maximize}")
                    continue

                sharpe_ratio = result[maximize]
                results.append({**flat_params, 'Sharpe Ratio': sharpe_ratio})
                print(f"Sharpe ratio: {sharpe_ratio}")

                if sharpe_ratio > best_sharpe:
                    best_sharpe = sharpe_ratio
                    best_params = params  # Store the full params dict
                    print(f"New best Sharpe ratio: {best_sharpe}")
                    
            except Exception as e:
                print(f"Error testing parameters {flat_params}: {e}")
                continue

        results_df = pd.DataFrame(results)
        
        # Print optimization summary
        if results:
            print(f"\n=== Optimization Complete ===")
            print(f"Tested {len(param_combinations)} parameter combinations")
            print(f"Valid results: {len(results)}")
            print(f"Best {maximize}: {best_sharpe:.4f}")
            print(f"Best parameters: {best_params}")
            
            if len(results) > 1:
                print(f"Worst {maximize}: {min(results, key=lambda x: x['Sharpe Ratio'])['Sharpe Ratio']:.4f}")
                print(f"Average {maximize}: {results_df['Sharpe Ratio'].mean():.4f}")
        else:
            print("Warning: No valid results found during optimization")

        return best_params, best_sharpe, results_df

    def _optimize_sambo(self, hyperparameter_grid: dict, maximize: str, start_date: str, end_date: str,
                       max_tries: int, constraint: callable, random_state: int, 
                       return_heatmap: bool, return_optimization: bool):
        """SAMBO (Bayesian) optimization implementation"""
        try:
            import sambo
        except ImportError:
            raise ImportError("Need package 'sambo' for method='sambo'. Install with: pip install sambo") from None

        # Set default max_tries for SAMBO
        if max_tries is None:
            max_tries = 10
        elif 0 < max_tries <= 1:
            # If max_tries is a fraction, convert to absolute number
            grid_size = np.prod([len(values) for values in hyperparameter_grid.values()])
            max_tries = max(1, int(max_tries * grid_size))
        
        print(f"SAMBO optimization with {max_tries} iterations")

        # Convert hyperparameter grid to SAMBO dimensions
        dimensions = []
        param_names = list(hyperparameter_grid.keys())
        
        for key, values in hyperparameter_grid.items():
            values = np.asarray(values)
            
            # Handle different data types
            if values.dtype.kind in 'mM':  # timedelta, datetime64
                # Convert to int64 for SAMBO compatibility
                values = values.astype(np.int64)

            if values.dtype.kind in 'iumM':  # integer types
                dimensions.append((values.min(), values.max() + 1))
            elif values.dtype.kind == 'f':  # float types
                dimensions.append((values.min(), values.max()))
            else:  # categorical
                dimensions.append(values.tolist())

        print(f"Parameter dimensions: {dimensions}")

        # Counter for tracking evaluations
        eval_count = [0]  # Use list to make it mutable in nested function
        
        # Memoized run function to avoid recomputing same parameters
        @lru_cache()
        def memoized_run(param_tuple):
            flat_params = dict(param_tuple)
            
            # Construct the full parameter dictionary
            params = {
                "iv_slope_thresholds": {
                    "upper_gamma": flat_params["upper_gamma"],
                    "upper_buffer": flat_params["upper_buffer"],
                    "lower_buffer": flat_params["lower_buffer"],
                    "lower_gamma": flat_params["lower_gamma"]
                },
                "portfolio_sl": flat_params.get("portfolio_sl", 0.01),
                "portfolio_tp": flat_params.get("portfolio_tp", 0.03),
                "legs": self.legs
            }
            
            try:
                backtest = self._backtest_factory()
                
                if start_date is not None or end_date is not None:
                    result = backtest.run_window(start_date=start_date, end_date=end_date, **params)
                else:
                    result = backtest.run(**params)
                
                if maximize not in result or pd.isna(result[maximize]):
                    print(f"Invalid result for params {flat_params}: {result}")
                    return 1000.0  # Return large penalty for invalid results
                
                # Return negative value since SAMBO minimizes but we want to maximize
                return -float(result[maximize])
                
            except Exception as e:
                print(f"Error in backtest with params {flat_params}: {e}")
                return 1000.0  # Return large penalty for failed backtests

        def objective_function(x):
            eval_count[0] += 1
            if eval_count[0] % 10 == 0:
                print(f"Evaluation {eval_count[0]}/{max_tries}")
            
            param_dict = dict(zip(param_names, x))
            param_tuple = tuple(param_dict.items())
            
            try:
                value = memoized_run(param_tuple)
                # Ensure we return a finite number
                if np.isnan(value) or np.isinf(value):
                    return 1000.0
                return float(value)
            except Exception as e:
                print(f"Error in objective function: {e}")
                return 1000.0

        # Simplified constraints function
        constraints = None
        if constraint is not None:
            def cons(x):
                try:
                    param_dict = dict(zip(param_names, x))
                    return constraint(param_dict)
                except Exception as e:
                    print(f"Error in constraint function: {e}")
                    return False
            constraints = cons

        try:
            print("Starting SAMBO optimization...")
            # Run SAMBO optimization
            res = sambo.minimize(
                fun=objective_function,
                bounds=dimensions,
                constraints=constraints,
                max_iter=max_tries,
                method='sceua',  # Shuffled Complex Evolution
                rng=random_state
            )
            print("SAMBO optimization completed.")
        
        except Exception as e:
            print(f"Error during SAMBO optimization: {e}")
            # Return best grid search result as fallback
            print("Falling back to grid search...")
            return self._optimize_grid(hyperparameter_grid, maximize, start_date, end_date, max_tries, constraint)

        # Get the best parameters and reconstruct full parameter dict
        best_flat_params = dict(zip(param_names, res.x))
        best_params = {
            "iv_slope_thresholds": {
                "upper_gamma": best_flat_params["upper_gamma"],
                "upper_buffer": best_flat_params["upper_buffer"],
                "lower_buffer": best_flat_params["lower_buffer"],
                "lower_gamma": best_flat_params["lower_gamma"]
            },
            "portfolio_sl": best_flat_params.get("portfolio_sl", 0.01),
            "portfolio_tp": best_flat_params.get("portfolio_tp", 0.03),
            "legs": self.legs
        }

        # Get the best score (convert back from negative)
        best_score = -res.fun
        
        print(f"\n=== SAMBO Optimization Complete ===")
        print(f"Total evaluations: {eval_count[0]}")
        print(f"Best {maximize}: {best_score:.4f}")
        print(f"Best parameters: {best_params}")

        # Prepare output
        # output = [best_params, best_score]

        # if return_heatmap:
        #     # Create heatmap from optimization history
        #     try:
        #         heatmap = pd.Series(dict(zip(map(tuple, res.xv), -res.funv)),
        #                           name=maximize)
        #         heatmap.index.names = param_names
        #         heatmap.sort_index(inplace=True)
        #         output.append(heatmap)
        #     except Exception as e:
        #         print(f"Could not create heatmap: {e}")
        #         output.append(None)

        # if return_optimization:
        #     output.append(res)
            
        return best_params, best_score

        # return tuple(output) if len(output) > 2 else (best_params, best_score)

    def generate_heatmaps(self, results_df: 'pd.DataFrame', output_file: str, png_directory: str) -> None:
        workbook = Workbook()
        sheet = workbook.create_sheet(title="Sharpe Ratio")
        for row_idx, row in enumerate([results_df.columns.tolist()] + results_df.values.tolist(), start=1):
            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        start_row = len(results_df) + 3
        parameter_columns = [col for col in results_df.columns if col != "Sharpe Ratio"]
        parameter_pairs = list(itertools.combinations(parameter_columns, 2))
        for param_x, param_y in parameter_pairs:
            grouped = results_df.groupby([param_x, param_y])["Sharpe Ratio"].mean().unstack(fill_value=np.nan)
            cmap = plt.cm.viridis.copy()
            cmap.set_bad(color="white")
            plt.figure(figsize=(5, 4))
            ax = plt.gca()
            heatmap = ax.imshow(grouped.values, cmap=cmap, aspect="auto", interpolation="nearest")
            ax.set_xticks(np.arange(len(grouped.columns)))
            ax.set_yticks(np.arange(len(grouped.index)))
            ax.set_xticklabels(grouped.columns, rotation=45)
            ax.set_yticklabels(grouped.index)
            ax.set_xlabel(param_y)
            ax.set_ylabel(param_x)
            ax.set_title(f"Heatmap of Sharpe Ratio: {param_x} vs {param_y}")
            plt.colorbar(heatmap, ax=ax)
            image_file = os.path.join(png_directory, f"{param_x}_{param_y}_Sharpe_Ratio_heatmap.png")
            plt.savefig(image_file, bbox_inches="tight", dpi=150)
            plt.close()
            from openpyxl.drawing.image import Image
            img = Image(image_file)
            img.anchor = f"A{start_row}"
            sheet.add_image(img)
            start_row += 30
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])
        workbook.save(output_file)
        print(f"Heatmaps saved to {output_file}")

    def delete_png_files(self, png_directory: str) -> None:
        for filename in os.listdir(png_directory):
            if filename.endswith(".png"):
                file_path = os.path.join(png_directory, filename)
                os.remove(file_path)