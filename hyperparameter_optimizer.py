import os
import re
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from itertools import product
import itertools
import datetime
from openpyxl.drawing.image import Image
import json
from functools import partial
# from iv_slope.Hyper_Risk_framework.trade_manager import TradeManager
from backtesting_opt1 import Backtest

class HyperParameterOptimizer:
    def __init__(self, db_path: str, strategy, *, cash: float = 100000, commission_per_contract: float = 0.65, 
                 option_multiplier: int = 75, legs: dict):
        self._backtest_factory = partial(
            Backtest, db_path=db_path, strategy=strategy, cash=cash,
            commission_per_contract=commission_per_contract, option_multiplier=option_multiplier
        )
        self.legs = legs

    def optimize(self, hyperparameter_grid: dict, maximize: str = 'Sharpe Ratio', 
                 method: str = 'grid', max_tries: int = None, constraint: callable = None):
        if method != 'grid':
            raise ValueError("Only 'grid' method is supported in this implementation")
        param_keys = list(hyperparameter_grid.keys())
        param_values = [hyperparameter_grid[key] for key in param_keys]
        param_combinations = [dict(zip(param_keys, combo)) for combo in product(*param_values)]

        if constraint:
            param_combinations = [params for params in param_combinations if constraint(params)]
        if max_tries and max_tries < len(param_combinations):
            param_combinations = param_combinations[:max_tries]
        if not param_combinations:
            raise ValueError("No admissible parameter combinations to test")
        results = []
        best_sharpe = -np.inf
        best_params = None

        for flat_params in param_combinations:
            print(f"Testing parameters: {flat_params}")

            # Construct the full parameter dictionary
            params = {
                "iv_slope_thresholds": {
                    "upper_gamma": flat_params["upper_gamma"],
                    "upper_buffer": flat_params["upper_buffer"],
                    "lower_buffer": flat_params["lower_buffer"],
                    "lower_gamma": flat_params["lower_gamma"]
                },
                "portfolio_sl": flat_params.get("portfolio_sl", 0.01),  # Default if not optimized
                "portfolio_tp": flat_params.get("portfolio_tp", 0.03),  # Default if not optimized
                "legs": self.legs  # Pass legs as a parameter
            }

            backtest = self._backtest_factory()
            result = backtest.optimize(**params)
            print(f"results : {result}")

            if maximize not in result or pd.isna(result[maximize]):
                print(f"Invalid result for parameters {flat_params}: {result}")
                continue

            sharpe_ratio = result[maximize]
            results.append({**flat_params, 'Sharpe Ratio': sharpe_ratio})
            print(f"Sharpe ratio: {sharpe_ratio}")

            if sharpe_ratio > best_sharpe:
                best_sharpe = sharpe_ratio
                best_params = params  # Store the full params dict

        results_df = pd.DataFrame(results)
        # timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        # output_directory_path = os.path.join(os.getcwd(), f'hyperparameter_optimizer_output_{timestamp}')
        # os.makedirs(output_directory_path, exist_ok=True)
        # sheet_path = os.path.join(output_directory_path, 'summary.xlsx')
        # self.generate_heatmaps(results_df, sheet_path, output_directory_path)
        # self.delete_png_files(output_directory_path)

        return best_params, best_sharpe, results_df

    def generate_heatmaps(self, results_df: 'pd.DataFrame', output_file: str, png_directory: str) -> None:
        workbook = Workbook()
        sheet = workbook.create_sheet(title="Sharpe Ratio")
        for row_idx, row in enumerate([results_df.columns.tolist()] + results_df.values.tolist(), start=1):
            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        start_row = len(results_df) + 3
        parameter_columns = [col for col in results_df.columns if col != "Sharpe Ratio"]
        parameter_pairs = list(itertools.combinations(parameter_columns, 2))
        for param_x, param_y in parameter_pairs:
            grouped = results_df.groupby([param_x, param_y])["Sharpe Ratio"].mean().unstack(fill_value=np.nan)
            cmap = plt.cm.viridis.copy()
            cmap.set_bad(color="white")
            plt.figure(figsize=(5, 4))
            ax = plt.gca()
            heatmap = ax.imshow(grouped.values, cmap=cmap, aspect="auto", interpolation="nearest")
            ax.set_xticks(np.arange(len(grouped.columns)))
            ax.set_yticks(np.arange(len(grouped.index)))
            ax.set_xticklabels(grouped.columns, rotation=45)
            ax.set_yticklabels(grouped.index)
            ax.set_xlabel(param_y)
            ax.set_ylabel(param_x)
            ax.set_title(f"Heatmap of Sharpe Ratio: {param_x} vs {param_y}")
            plt.colorbar(heatmap, ax=ax)
            image_file = os.path.join(png_directory, f"{param_x}_{param_y}_Sharpe_Ratio_heatmap.png")
            plt.savefig(image_file, bbox_inches="tight", dpi=150)
            plt.close()
            from openpyxl.drawing.image import Image
            img = Image(image_file)
            img.anchor = f"A{start_row}"
            sheet.add_image(img)
            start_row += 30
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])
        workbook.save(output_file)
        print(f"Heatmaps saved to {output_file}")

    def delete_png_files(self, png_directory: str) -> None:
        for filename in os.listdir(png_directory):
            if filename.endswith(".png"):
                file_path = os.path.join(png_directory, filename)
                os.remove(file_path)